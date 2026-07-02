# -*- coding: utf-8 -*-
"""Парсер xlsx-отчёта института по хвостам.

Устройство листа «Итог» (раздел 3 CLAUDE.md):
1) баланс: Шихта руд / Материал пруда / Итого / Отвальные хвосты;
2) на каждую разновидность хвостов (породные / пирротиновые / общие) — секция:
   строка баланса «Хвосты …» -> таблица «Класс крупности, мкм» -> блоки
   минералогии по классам -> «Итого извлекаемый / не извлекаемый металл».
Метки «гуляют» между файлами — все метки нормализуются (labels.py).
Битые значения (#REF!, текст) не роняют парсер: значение None + issue,
восстановлением занимается parser/recover.py.
Пустая ячейка в таблице данных = 0 (так институт помечает нули) — measured.
"""
from __future__ import annotations

import io
import re
from pathlib import Path

import openpyxl
from pydantic import BaseModel, Field

from ..models import (RECOVERABLE, DataIssue, LossCell, MineralForm,
                      SizeClassRow, TailingsReport)
from .labels import (SPECIAL_FREE_SLOT, SPECIAL_UNALLOCATED, norm_text,
                     normalize_form_label, normalize_size_label)


class ParseResult(BaseModel):
    source: str = ""
    plant: str = ""
    reports: list[TailingsReport] = Field(default_factory=list)
    issues: list[DataIssue] = Field(default_factory=list)   # файл-уровневые
    meta: dict = Field(default_factory=dict)                 # Факт/Расчёт, Поступило в переработку

    @property
    def all_issues(self) -> list[DataIssue]:
        out = list(self.issues)
        for r in self.reports:
            out.extend(r.issues)
        return out


def _parse_value(v) -> tuple[str, float | None, str]:
    """-> (kind, value, raw); kind: num | blank | broken | text."""
    if v is None:
        return "blank", None, ""
    if isinstance(v, (int, float)):
        return "num", float(v), str(v)
    s = str(v).strip()
    if not s:
        return "blank", None, ""
    if s.startswith("#"):  # #REF!, #Н/Д, #DIV/0! ...
        return "broken", None, s
    try:
        return "num", float(s.replace(" ", "").replace(",", ".").replace(" ", "")), s
    except ValueError:
        return "text", None, s


def detect_plant(source_name: str) -> str:
    s = norm_text(source_name)
    low = s.lower()
    if "кгмк" in low:
        return "КГМК"
    if "ноф" in low:
        return "НОФ (медистые)" if "мед" in low else "НОФ (вкрапленные)"
    if "тоф" in low:
        return "ТОФ"
    return s or "Фабрика"


class _Row:
    __slots__ = ("idx", "label", "cells")

    def __init__(self, idx: int, label: str, cells: list):
        self.idx = idx          # номер строки excel
        self.label = label      # NFC-нормализованная метка (из колонок A..C)
        self.cells = cells      # 5 значений C..G: список (kind, value, coord)


def parse_workbook(source, source_name: str = "") -> ParseResult:
    """source: путь | bytes | file-like. Никогда не бросает на битых данных."""
    if isinstance(source, (str, Path)):
        wb = openpyxl.load_workbook(source, data_only=True)
        source_name = source_name or Path(source).name
    elif isinstance(source, bytes):
        wb = openpyxl.load_workbook(io.BytesIO(source), data_only=True)
    else:
        wb = openpyxl.load_workbook(source, data_only=True)

    result = ParseResult(source=norm_text(source_name), plant=detect_plant(source_name))

    sheet = None
    for sn in wb.sheetnames:
        if norm_text(sn).lower() == "итог":
            sheet = wb[sn]
            break
    if sheet is None:
        sheet = wb[wb.sheetnames[0]]
        result.issues.append(DataIssue(
            severity="warning", rule="parser",
            message=f"Лист «Итог» не найден, использую первый лист «{sheet.title}»"))

    rows = _read_rows(sheet)
    _parse_global(rows, result)

    header_idxs = [i for i, r in enumerate(rows)
                   if r.label and "класс крупности" in r.label.lower()]
    if not header_idxs:
        result.issues.append(DataIssue(
            severity="error", rule="parser",
            message="Не найдена ни одна таблица «Класс крупности» — структура файла неизвестна"))
        return result

    for n, hi in enumerate(header_idxs):
        end = header_idxs[n + 1] - 3 if n + 1 < len(header_idxs) else len(rows)
        report = _parse_section(rows, hi, end, result)
        if report:
            result.reports.append(report)
    return result


def _read_rows(ws) -> list[_Row]:
    rows: list[_Row] = []
    for excel_row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=8):
        label = ""
        for c in excel_row[:3]:  # A..C
            if isinstance(c.value, str) and c.value.strip():
                label = norm_text(c.value)
                break
        # значения C..G (индексы 2..6)
        cells = []
        for c in excel_row[2:7]:
            kind, val, raw = _parse_value(c.value)
            cells.append((kind, val, c.coordinate, raw))
        rows.append(_Row(excel_row[0].row, label, cells))
    return rows


def _num_or_none(cell, report: TailingsReport | None, where: str,
                 blank_is_zero: bool = True) -> float | None:
    """Значение ячейки; broken/text -> None + issue."""
    kind, val, coord, _raw = cell
    if kind == "num":
        return val
    if kind == "blank":
        return 0.0 if blank_is_zero else None
    issue = DataIssue(
        severity="error", rule="parser", cell=where,
        message=f"Битое значение в {coord} ({where}): "
                f"{'#REF!/ошибка формулы' if kind == 'broken' else 'не число'} — требует восстановления")
    if report is not None:
        report.issues.append(issue)
    return None


def _parse_global(rows: list[_Row], result: ParseResult):
    """Баланс «Поступило в переработку» и общие строки «Отвальные хвосты» (Факт/Расчёт)."""
    mode = None
    for i, r in enumerate(rows):
        if i > 40:
            break
        low = (r.label or "").lower()
        if low == "факт":
            mode = "Факт"
        elif low.startswith("расч"):
            mode = "Расчёт"
        elif low == "итого" and "feed" not in result.meta:
            vals = [c[1] if c[0] == "num" else None for c in r.cells]
            result.meta["feed"] = {"смт": vals[0], "Ni_pct": vals[1], "Ni_t": vals[2],
                                   "Cu_pct": vals[3], "Cu_t": vals[4]}
        elif low.startswith("отвальные хвосты"):
            vals = [c[1] if c[0] == "num" else None for c in r.cells]
            key = f"отвальные_{mode or 'Факт'}"
            result.meta.setdefault(key, {"смт": vals[0], "Ni_pct": vals[1], "Ni_t": vals[2],
                                         "Cu_pct": vals[3], "Cu_t": vals[4]})


_BALANCE_RE = re.compile(r"^(отвальные\s+)?хвосты", re.IGNORECASE)


def _parse_section(rows: list[_Row], hi: int, end: int, result: ParseResult) -> TailingsReport | None:
    # 1) строка баланса секции — выше заголовка таблицы крупности
    balance = None
    for j in range(hi - 1, max(hi - 8, -1), -1):
        r = rows[j]
        if r.label and _BALANCE_RE.match(r.label) and any(c[0] == "num" for c in r.cells):
            balance = r
            break

    low_label = (balance.label if balance else "").lower()
    if "пирротин" in low_label:
        tail_type = "пирротиновые"
    elif "породн" in low_label:
        tail_type = "породные"
    else:
        tail_type = "отвальные (общие)"

    report = TailingsReport(plant=result.plant, tail_type=tail_type)

    if balance:
        report.tails_tonnes = _num_or_none(balance.cells[0], report, f"{tail_type}: СМТ")
        ni_pct = _num_or_none(balance.cells[1], report, f"{tail_type}: Ni %")
        ni_t = _num_or_none(balance.cells[2], report, f"{tail_type}: Ni т")
        cu_pct = _num_or_none(balance.cells[3], report, f"{tail_type}: Cu %")
        cu_t = _num_or_none(balance.cells[4], report, f"{tail_type}: Cu т")
        report.grade = {"Ni": ni_pct, "Cu": cu_pct}
        report.losses_tonnes = {"Ni": ni_t, "Cu": cu_t}
    else:
        report.issues.append(DataIssue(
            severity="warning", rule="parser",
            message="Не найдена строка баланса секции («Хвосты …») перед таблицей крупности"))
    feed = result.meta.get("feed") or {}
    report.feed_tonnes = feed.get("смт")

    # 2) таблица крупности
    i = hi + 1
    while i < end:
        r = rows[i]
        if r.label:
            if r.label.lower().startswith("итого"):
                report.control_totals["size_total"] = _row_values(r, report, "Итого таблицы крупности")
                i += 1
                break
            canon = normalize_size_label(r.label)
            if canon:
                report.size_classes.append(SizeClassRow(
                    label=canon,
                    share_pct=_num_or_none(r.cells[0], report, f"{canon}: доля класса %"),
                    element_share_pct={
                        "Ni": _num_or_none(r.cells[1], report, f"{canon}: доля Ni %"),
                        "Cu": _num_or_none(r.cells[3], report, f"{canon}: доля Cu %"),
                    },
                    element_tonnes={
                        "Ni": _num_or_none(r.cells[2], report, f"{canon}: Ni т"),
                        "Cu": _num_or_none(r.cells[4], report, f"{canon}: Cu т"),
                    },
                ))
        i += 1

    # 3) блоки минералогии: заголовок = метка класса + «Доля потерь …» в D
    while i < end:
        r = rows[i]
        if _is_mineral_header(r):
            i = _parse_mineral_block(rows, i, end, report)
            continue
        low = (r.label or "").lower()
        if low.startswith("итого извлекаемый"):
            v = _row_values(r, report, "Итого извлекаемый металл")
            report.recoverable_pct = {"Ni": v.get("Ni_share"), "Cu": v.get("Cu_share")}
            report.recoverable_total = {"Ni": v.get("Ni_t"), "Cu": v.get("Cu_t")}
        elif low.startswith("итого не извлекаемый"):
            report.control_totals["nonrecoverable"] = _row_values(r, report, "Итого не извлекаемый металл")
        elif low.startswith("итого"):
            report.control_totals["grand_check"] = _row_values(r, report, "Итого (проверка) секции")
        i += 1

    if not report.size_classes:
        report.issues.append(DataIssue(
            severity="error", rule="parser",
            message=f"Секция «{tail_type}»: таблица крупности пуста"))
        return None
    return report


def _row_values(r: _Row, report: TailingsReport | None, where: str) -> dict:
    return {
        "share": _num_or_none(r.cells[0], None, where, blank_is_zero=False),
        "Ni_share": _num_or_none(r.cells[1], None, where, blank_is_zero=False),
        "Ni_t": _num_or_none(r.cells[2], None, where, blank_is_zero=False),
        "Cu_share": _num_or_none(r.cells[3], None, where, blank_is_zero=False),
        "Cu_t": _num_or_none(r.cells[4], None, where, blank_is_zero=False),
    }


def _is_mineral_header(r: _Row) -> bool:
    if not r.label:
        return False
    kind, _val, _coord, raw = r.cells[1]
    # надёжный признак заголовка блока: в колонке D строка «Доля потерь …»
    if kind != "text" or "доля потерь" not in raw.lower():
        return False
    return normalize_size_label(r.label) is not None


def _parse_mineral_block(rows: list[_Row], hi: int, end: int, report: TailingsReport) -> int:
    size_class = normalize_size_label(rows[hi].label)
    i = hi + 1
    seen_forms: set[str] = set()
    while i < end:
        r = rows[i]
        low = (r.label or "").lower()
        if not r.label:
            i += 1
            continue
        if low.startswith("итого"):
            report.control_totals.setdefault("class_totals", {})[size_class] = \
                _row_values(r, report, f"{size_class}: Итого (проверка)")
            i += 1
            break
        if _is_mineral_header(r):  # защита: следующий блок начался без «Итого»
            return i
        form = normalize_form_label(r.label)
        if form in (SPECIAL_UNALLOCATED, SPECIAL_FREE_SLOT):
            for col, el in ((2, "Ni"), (4, "Cu")):
                kind, val, coord, _raw = r.cells[col]
                if kind == "num" and val and abs(val) > 1e-9:
                    report.issues.append(DataIssue(
                        severity="warning", rule="parser",
                        cell=f"{size_class} / {r.label} / {el}",
                        message=f"Нераспределённые потери {val:.2f} т ({el}, класс {size_class}) — "
                                f"проверьте расшифровку анализа"))
            i += 1
            continue
        if form is None:
            report.issues.append(DataIssue(
                severity="info", rule="parser", cell=f"{size_class} / {r.label}",
                message=f"Неизвестная строка минералогии «{r.label}» (класс {size_class}) — пропущена"))
            i += 1
            continue
        if form in seen_forms:
            i += 1
            continue
        seen_forms.add(form)
        for share_col, tonnes_col, el in ((1, 2, "Ni"), (3, 4, "Cu")):
            where = f"{size_class} / {form} / {el}"
            report.cells.append(LossCell(
                axes={"size_class": size_class, "mineral_form": form},
                element=el,
                share_pct=_num_or_none(r.cells[share_col], report, where + " (доля %)"),
                tonnes=_num_or_none(r.cells[tonnes_col], report, where + " (т)"),
                recoverable=MineralForm(form) in RECOVERABLE[el],
            ))
        i += 1

    # строки «Извлекаемый металл» / «Не извлекаемый металл» после Итого
    scanned = 0
    while i < end and scanned < 5:
        r = rows[i]
        low = (r.label or "").lower()
        if low.startswith("извлекаемый"):
            report.control_totals.setdefault("class_recoverable", {})[size_class] = \
                _row_values(r, report, f"{size_class}: извлекаемый металл")
        elif low.startswith("не извлекаемый"):
            report.control_totals.setdefault("class_nonrecoverable", {})[size_class] = \
                _row_values(r, report, f"{size_class}: не извлекаемый металл")
        elif low:
            break  # «Итого извлекаемый металл» и прочее — отдаём секции
        i += 1
        scanned += 1
    return i
