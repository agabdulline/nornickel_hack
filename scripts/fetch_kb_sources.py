# -*- coding: utf-8 -*-
"""Скачивание внешних источников базы знаний по kb_extra_sources.xlsx.

- КиберЛенинка: пробуем PDF (<url>/pdf); капча/блок -> текст статьи из HTML
  (div.ocr), первой строкой «Автор(ы). Год. Название».
- Патенты (findpatent/patenton): HTML -> полный текст, первой строкой
  «Патент RU <номер>. <Название>. <Год>».
- Веб-статьи: текст с шапкой.
Результат: data/kb/extra/* + data/kb/sources.csv (те же колонки + status).
"""
from __future__ import annotations

import csv
import re
import sys
import time
from pathlib import Path

import httpx
import openpyxl
from bs4 import BeautifulSoup

ROOT = Path(__file__).resolve().parents[1]
XLSX = ROOT / "kb_extra_sources.xlsx"
OUT = ROOT / "data" / "kb" / "extra"
CSV_OUT = ROOT / "data" / "kb" / "sources.csv"

HEADERS = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
                  "(KHTML, like Gecko) Chrome/126.0.0.0 Safari/537.36",
    "Accept-Language": "ru-RU,ru;q=0.9,en;q=0.5",
}
ATTEMPTS = 2
TIMEOUT = 30.0


def load_rows() -> list[dict]:
    wb = openpyxl.load_workbook(XLSX)
    ws = wb[wb.sheetnames[0]]
    headers = [c.value for c in next(ws.iter_rows(max_row=1))]
    rows = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        if not any(r):
            continue
        rows.append({h: ("" if v is None else str(v).strip()) for h, v in zip(headers, r)})
    return fix_defects(rows)


def fix_defects(rows: list[dict]) -> list[dict]:
    """Строка Derrick задублирована/склеена (имя файла + кусок автора через таб)."""
    fixed = []
    for row in rows:
        f = row["file"]
        if "derrick" in f.lower():
            row = {**row,
                   "file": "2006_stroyka_tonkoe_grohochenie_derrick.txt",
                   "author": "Журнал «Стройка» (СПб) — материал о технологии Derrick",
                   "year": "2006",
                   "title": "Тонкое грохочение в технологии обогащения минерального сырья",
                   "type": "web"}
            if any("derrick" in x["file"].lower() for x in fixed):
                continue  # дубль
        row["year"] = row["year"].replace(".0", "")
        fixed.append(row)
    return fixed


def get(client: httpx.Client, url: str) -> httpx.Response | None:
    for attempt in range(ATTEMPTS):
        try:
            resp = client.get(url, timeout=TIMEOUT, follow_redirects=True)
            if resp.status_code == 200:
                return resp
            print(f"    HTTP {resp.status_code} ({url})")
        except httpx.HTTPError as e:
            print(f"    попытка {attempt + 1}: {type(e).__name__}")
        time.sleep(2 * (attempt + 1))
    return None


def html_to_text(html: bytes | str, base_url: str) -> str:
    soup = BeautifulSoup(html, "lxml")
    for tag in soup(["script", "style", "nav", "header", "footer", "form", "noscript",
                     "iframe", "button", "aside"]):
        tag.decompose()
    # КиберЛенинка: полный текст лежит в div.ocr
    if "cyberleninka" in base_url:
        ocr = soup.select("div.ocr p, div[itemprop=articleBody] p")
        if ocr:
            return "\n".join(p.get_text(" ", strip=True) for p in ocr)
    main = soup.find("main") or soup.find("article") or soup.body or soup
    text = main.get_text("\n", strip=True)
    text = re.sub(r"\n{3,}", "\n\n", text)
    return text


def looks_like_captcha(text: str) -> bool:
    low = text.lower()
    return len(text) < 3000 and any(m in low for m in
                                    ("captcha", "капч", "robot", "доступ ограничен", "cloudflare"))


def header_line(row: dict) -> str:
    if row["type"] == "patent":
        m = re.search(r"(\d{7})", row["file"])
        num = m.group(1) if m else ""
        return f"Патент RU {num}. {row['title']}. {row['year']}"
    return f"{row['author']}. {row['year']}. {row['title']}"


def google_patent_text(row: dict, client: httpx.Client) -> str | None:
    """Фоллбэк: то же описание патента с patents.google.com."""
    m = re.search(r"(\d{7})", row["file"])
    if not m:
        return None
    num = m.group(1)
    for kind in ("C1", "C2", ""):
        resp = get(client, f"https://patents.google.com/patent/RU{num}{kind}/ru")
        if resp is None:
            continue
        text = html_to_text(resp.content, "patents.google.com")
        if len(text) > 3000:
            print(f"    взято с patents.google.com (RU{num}{kind})")
            return text
    return None


def pdf_text_ok(path: Path) -> bool:
    import fitz
    with fitz.open(path) as doc:
        n = doc.page_count
        idxs = sorted({min(n - 1, round(i * (n - 1) / 4)) for i in range(min(5, n))})
        total = sum(len(doc[i].get_text() or "") for i in idxs)
        return total / max(len(idxs), 1) > 200


def process(row: dict, client: httpx.Client) -> dict:
    fname = row["file"]
    target = OUT / fname
    url = row.get("download_link") or row["url"]
    print(f"[{fname}]")
    status, text_ok = "failed", False

    if fname.endswith(".pdf"):
        resp = get(client, url)
        if resp and resp.content[:5] == b"%PDF-":
            target.write_bytes(resp.content)
            text_ok = pdf_text_ok(target)
            status = "ok_pdf"
        elif resp:
            print("    не PDF в ответе")
    else:
        text = None
        if "cyberleninka" in url:
            # в таблице download_link может уже оканчиваться на /pdf
            article_url = re.sub(r"/pdf/?$", "", url)
            pdf_resp = get(client, article_url + "/pdf")
            if pdf_resp is not None and pdf_resp.content[:5] == b"%PDF-":
                pdf_target = target.with_suffix(".pdf")
                pdf_target.write_bytes(pdf_resp.content)
                if pdf_text_ok(pdf_target):
                    row = {**row, "file": pdf_target.name}
                    print(f"    PDF ок -> {pdf_target.name}")
                    return {**row, "status": "ok_pdf", "text_layer_ok": "yes"}
                pdf_target.unlink()
            url = article_url  # HTML-фоллбэк — со страницы статьи, не /pdf
        resp = get(client, url)
        if resp is not None:
            text = html_to_text(resp.content, url)
            if looks_like_captcha(text):
                print("    похоже на капчу/блок")
                text = None
        if (not text or len(text) <= 2000) and row["type"] == "patent":
            text = google_patent_text(row, client) or text
        if text and len(text) > 2000:
            target.write_text(header_line(row) + "\n\n" + text, encoding="utf-8")
            status, text_ok = "ok_txt", True
        elif text:
            print(f"    текст подозрительно короткий ({len(text)} симв.) — не сохраняю")

    return {**row, "status": status, "text_layer_ok": "yes" if text_ok else "no"}


def main():
    OUT.mkdir(parents=True, exist_ok=True)
    rows = load_rows()
    print(f"источников в таблице после чистки: {len(rows)}")
    results = []
    with httpx.Client(headers=HEADERS) as client:
        for row in rows:
            results.append(process(row, client))
            time.sleep(1)

    with CSV_OUT.open("w", encoding="utf-8-sig", newline="") as f:
        w = csv.DictWriter(f, fieldnames=list(results[0].keys()), delimiter=";")
        w.writeheader()
        w.writerows(results)

    ok = [r for r in results if r["status"].startswith("ok")]
    print(f"\nскачано {len(ok)}/{len(results)} -> {OUT}")
    for r in results:
        print(f"  {r['status']:7} text_ok={r['text_layer_ok']:3} {r['file']}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
